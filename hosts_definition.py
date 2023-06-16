#!/usr/bin/env python3

import argparse
from dataclasses import dataclass
import ipaddress
import logging
from pprint import pformat
import re
from typing import List

import macaddress
import openpyxl

DEFAULT_WORKSHEET_NAME = 'hosts-definition'
OUTPUT_TYPES = ['dns', 'dhcp']
HOSTS_DEFINITION_HEADERS = ['hostname', 'domain', 'ip-address', 'mac-address', 'mx-preference']
DEFAULT_TTL = '1h'

DNS_LABEL = re.compile(r'[a-z0-9]([a-z0-9-]{0,61}[a-z0-9])?', re.IGNORECASE)
DNS_HOSTNAME = re.compile(rf'^({DNS_LABEL.pattern})$', re.IGNORECASE)
DNS_DOMAIN = re.compile(rf'^({DNS_LABEL.pattern})(\.{DNS_LABEL.pattern})+$', re.IGNORECASE)
MX_PREFERENCE = re.compile(r'^(\d{1,2})$')


@dataclass
class HostDefinition:
    hostname: str
    domain: str
    ip_address: ipaddress.IPv4Address = None
    mac_address: macaddress.HWAddress = None
    mx_preference: int = None


def args_parser():
    parser = argparse.ArgumentParser(description='Script to generate hosts definition for Mikrotik router')

    parser.add_argument('-v', '--verbose',
                        action='store_true',
                        default=False,
                        help='talkative mode')

    parser.add_argument('-f', '--file',
                        action='store',
                        required=True,
                        help='xlsx file with hosts definition')

    parser.add_argument('-o', '--output',
                        action='store',
                        help='output file')

    parser.add_argument('-t', '--output-type',
                        action='store',
                        choices=OUTPUT_TYPES,
                        required=True,
                        help=f"type of output, to be selected from {OUTPUT_TYPES}")

    params = parser.parse_args()

    return params


def logging_setup(verbose: bool):
    if verbose:
        logging.basicConfig(level=logging.DEBUG)
    else:
        logging.basicConfig(level=logging.INFO)


def read_hosts_definition_file(input_file_name: str) -> List[HostDefinition]:
    ret_definition = []

    workbook = openpyxl.load_workbook(filename=input_file_name, read_only=True)
    worksheet = workbook[DEFAULT_WORKSHEET_NAME]

    table_start_row = 2
    table_end_row = worksheet.max_row
    table_start_column = 1
    table_end_column = worksheet.max_column

    for row in worksheet.iter_rows(min_row=table_start_row,
                                   max_row=table_end_row,
                                   min_col=table_start_column,
                                   max_col=table_end_column):

        if not DNS_HOSTNAME.match(row[0].value):
            raise ValueError
        if not DNS_DOMAIN.match(row[1].value):
            raise ValueError
        if not ipaddress.ip_address(row[2].value):
            raise
        if row[3].value and not macaddress.MAC(row[3].value):
            raise
        if row[4].value and not MX_PREFERENCE.match(str(row[4].value)):
            raise ValueError

        hd = HostDefinition(hostname=row[0].value,
                            domain=row[1].value,
                            ip_address=row[2].value,
                            mac_address=row[3].value,
                            mx_preference=row[4].value)
        ret_definition.append(hd)

    workbook.close()

    logging.debug("row: %s", pformat(ret_definition))

    return ret_definition


def generate_dns_output(definition, output_file: str = None) -> None:
    output_lines = []

    for item in definition:
        output_lines.append(f"add name={item.hostname}.{item.domain} "
                            f"address={item.ip_address} "
                            f"ttl={DEFAULT_TTL}")

        if item.mx_preference:
            output_lines.append(f"add name={item.domain} "
                                f"type=MX "
                                f"mx-exchange={item.hostname}.{item.domain} "
                                f"mx-preference={item.mx_preference} "
                                f"ttl={DEFAULT_TTL}")

    write_output(output_file, output_lines)


def generate_dhcp_output(definition, output_file: str = None) -> None:
    output_lines = []

    for item in [item for item in definition if item.mac_address]:
        output_lines.append(f"add mac-address={item.mac_address} "
                            f"address={item.ip_address} "
                            f"comment={item.hostname}.{item.domain}")

    write_output(output_file, output_lines)


def write_output(output_file: str, output_lines: List[str]):
    if output_file:
        with open(output_file, mode='w', encoding='utf-8') as f:
            f.write('\n'.join(output_lines))
    else:
        print('\n'.join(output_lines))


def main():
    params = args_parser()

    logging_setup(params.verbose)

    logging.debug(params)

    definition = read_hosts_definition_file(params.file)

    if params.output_type == 'dns':
        generate_dns_output(definition, params.output)
    elif params.output_type == 'dhcp':
        generate_dhcp_output(definition)


if __name__ == '__main__':
    main()
